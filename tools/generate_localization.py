#!/usr/bin/env python3
"""
Generates Honeycomb's localization string tables from a single source of
truth: the Honeycomb_Localization.xlsx spreadsheet (Section / Key / Context /
English / Spanish / Translate columns).

Source of truth: Honeycomb_Localization.xlsx lives in this repo at
tools/Honeycomb_Localization.xlsx (committed to git) — edit that copy
directly and re-run this script with no arguments. Never hand-edit any of
the generated output files below; they're regenerated wholesale each run.

Usage:
    python3 tools/generate_localization.py [path/to/Honeycomb_Localization.xlsx]

Outputs:
    shared/Localization/StringKey.swift             (mac + iOS)
    shared/Localization/Strings.English.swift        (mac + iOS)
    shared/Localization/Strings.Spanish.swift         (mac + iOS)
    windows/src/SoliBee.Core/Localization/StringKey.cs
    windows/src/SoliBee.Core/Localization/Strings.cs

Key rules:
  - `Key` is the spreadsheet's own snake_case identifier (author-assigned,
    not auto-derived from the English text) — this keeps enum case names
    stable and readable even if the English wording later changes.
  - `Translate` = FALSE means the Spanish column is intentionally identical
    to English (proper nouns, debug-only strings, script-name labels like
    "English"/"Español" that stay in their own language regardless of
    which language is active) — the row is still emitted, just with
    Spanish == English on purpose, not a leftover TODO.
  - Every row must have non-empty English AND Spanish text — this script
    fails loudly on a blank cell rather than emitting a silently-missing
    translation (English-fallback for a truly missing key is a runtime
    concern handled by the L()/Strings.Get() lookup, not something this
    generator should paper over at build time).
"""
import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
_REPO_COPY = REPO_ROOT / "tools" / "Honeycomb_Localization.xlsx"
_DOWNLOADS_COPY = Path.home() / "Downloads" / "Honeycomb_Localization.xlsx"
DEFAULT_SOURCE = _REPO_COPY if _REPO_COPY.exists() else _DOWNLOADS_COPY

OUT_SWIFT_KEY = REPO_ROOT / "shared/Localization/StringKey.swift"
OUT_SWIFT_EN = REPO_ROOT / "shared/Localization/Strings.English.swift"
OUT_SWIFT_ES = REPO_ROOT / "shared/Localization/Strings.Spanish.swift"
OUT_CS_KEY = REPO_ROOT / "windows/src/SoliBee.Core/Localization/StringKey.cs"
OUT_CS_STRINGS = REPO_ROOT / "windows/src/SoliBee.Core/Localization/Strings.cs"

EXPECTED_HEADERS = ["Section", "Key", "Context", "English", "Spanish", "Translate"]

_KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")


def snake_to_lower_camel(snake: str) -> str:
    parts = snake.split("_")
    return parts[0] + "".join(p.capitalize() for p in parts[1:])


def snake_to_pascal(snake: str) -> str:
    return "".join(p.capitalize() for p in snake.split("_"))


def swift_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")


def cs_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")


def load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected header row {headers!r}; expected {EXPECTED_HEADERS!r}")

    rows = []
    seen_keys: set[str] = set()
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None for v in vals):
            continue
        section, key, context, english, spanish, translate = vals

        missing = [name for name, v in zip(EXPECTED_HEADERS, vals) if v is None and name != "Translate"]
        if missing:
            raise ValueError(f"Row {r} is missing {missing}: {vals!r}")

        key = str(key).strip()
        if not _KEY_RE.match(key):
            raise ValueError(f"Row {r}: Key {key!r} must be snake_case (a-z, 0-9, _), starting with a letter")
        if key in seen_keys:
            raise ValueError(f"Row {r}: duplicate Key {key!r}")
        seen_keys.add(key)

        english = str(english)
        spanish = str(spanish)
        if not english.strip() or not spanish.strip():
            raise ValueError(f"Row {r} ({key}): English and Spanish must both be non-empty")

        rows.append({
            "section": str(section).strip(),
            "key": key,
            "context": str(context).strip() if context else "",
            "english": english,
            "spanish": spanish,
            "translate": bool(translate) if translate is not None else True,
        })
    return rows


def write_swift_key(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_localization.py` from",
        "// Honeycomb_Localization.xlsx.",
        "",
        "// Stable identifier for every localizable string (mac + iOS). Mirrors the",
        "// Windows port's StringKey (windows/src/SoliBee.Core/Localization/StringKey.cs)",
        "// — same keys, generated from the same spreadsheet in one pass.",
        "public enum StringKey: String, CaseIterable {",
    ]
    for row in rows:
        case_name = snake_to_lower_camel(row["key"])
        if row["context"]:
            lines.append(f"    /// {row['context']}")
        lines.append(f'    case {case_name} = "{row["key"]}"')
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_swift_table(rows: list[dict], path: Path, type_name: str, text_field: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_localization.py` from",
        "// Honeycomb_Localization.xlsx.",
        "",
        f"enum {type_name} {{",
        "    static let table: [StringKey: String] = [",
    ]
    for row in rows:
        case_name = snake_to_lower_camel(row["key"])
        text = swift_escape(row[text_field])
        lines.append(f'        .{case_name}: "{text}",')
    lines.append("    ]")
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cs_key(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_localization.py` from",
        "// Honeycomb_Localization.xlsx.",
        "",
        "namespace SoliBee.Core.Localization;",
        "",
        "// Stable identifier for every localizable string (Windows). Mirrors the",
        "// Mac/iOS port's StringKey (shared/Localization/StringKey.swift) — same",
        "// keys, generated from the same spreadsheet in one pass.",
        "public enum StringKey",
        "{",
    ]
    for row in rows:
        lines.append(f"    {snake_to_pascal(row['key'])},")
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cs_strings(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_localization.py` from",
        "// Honeycomb_Localization.xlsx.",
        "",
        "using System.Collections.Generic;",
        "",
        "namespace SoliBee.Core.Localization;",
        "",
        "// Falls back to English on a missing key (e.g. a newly-added string whose",
        "// Spanish translation hasn't landed in the spreadsheet yet) rather than",
        "// throwing — a missing translation should degrade, not crash the app.",
        "public static class Strings",
        "{",
        "    public static string Get(StringKey key, AppLanguage language)",
        "    {",
        "        var table = language == AppLanguage.Spanish ? Spanish : English;",
        "        if (table.TryGetValue(key, out var value)) return value;",
        "        return English.TryGetValue(key, out var fallback) ? fallback : $\"?{key}?\";",
        "    }",
        "",
        "    private static readonly Dictionary<StringKey, string> English = new()",
        "    {",
    ]
    for row in rows:
        lines.append(f'        [StringKey.{snake_to_pascal(row["key"])}] = "{cs_escape(row["english"])}",')
    lines.append("    };")
    lines.append("")
    lines.append("    private static readonly Dictionary<StringKey, string> Spanish = new()")
    lines.append("    {")
    for row in rows:
        lines.append(f'        [StringKey.{snake_to_pascal(row["key"])}] = "{cs_escape(row["spanish"])}",')
    lines.append("    };")
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        print(f"Source spreadsheet not found: {source}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(source)

    write_swift_key(rows, OUT_SWIFT_KEY)
    write_swift_table(rows, OUT_SWIFT_EN, "StringsEnglish", "english")
    write_swift_table(rows, OUT_SWIFT_ES, "StringsSpanish", "spanish")
    write_cs_key(rows, OUT_CS_KEY)
    write_cs_strings(rows, OUT_CS_STRINGS)

    untranslated = sum(1 for r in rows if not r["translate"])
    print(f"Read {len(rows)} spreadsheet rows ({untranslated} marked Translate=FALSE, "
          "kept identical by design).")
    for out in (OUT_SWIFT_KEY, OUT_SWIFT_EN, OUT_SWIFT_ES, OUT_CS_KEY, OUT_CS_STRINGS):
        print(f"Wrote {out.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
