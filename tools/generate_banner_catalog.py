#!/usr/bin/env python3
"""
Generates the Honeycomb banner/toast catalog and its matching Swift + C#
identifier enums from a single source of truth: the content spreadsheet.

Source of truth: Honeycomb_Fun_Messages.xlsx (Category / Trigger / Message /
Type / Location columns). Re-run this script after every spreadsheet edit.
Never hand-edit any of the generated output files below — they're
regenerated wholesale each run, so hand edits are silently discarded.

Usage:
    python3 tools/generate_banner_catalog.py [path/to/Honeycomb_Fun_Messages.xlsx]

The spreadsheet itself lives in this repo at tools/Honeycomb_Fun_Messages.xlsx
(committed to git) — edit that copy and re-run this script with no arguments.
It's the canonical copy; a Downloads copy is only ever a staging/editing
convenience, never the thing to rely on long-term.

Outputs:
    shared/Honeycomb/Resources/HoneycombBannerCatalog.json   (shared content — bundled by mac + Windows)
    shared/Honeycomb/Models/BannerID.swift                    (mac + iOS)
    windows/src/SoliBee.Core/Models/BannerId.cs                (Windows)

--- The rules this script encodes (resolved via product discussion) ---

Every catalog entry groups all spreadsheet rows sharing the exact same
(Category, Trigger) pair into one `messages` array — a trigger with several
phrasings is one entry with several strings, picked uniform-random at
fire-time, not several entries.

`gated` (20% chance) applies ONLY when a trigger is standing in for an
existing production banner that would otherwise show:
  - Every `Game intro rules banner` row is gated with a dynamic fallback:
    80% of the time the rule's own existing display name shows instead
    (fallback: "$RULE_NAME", resolved by the caller at runtime — not a
    literal string, since it depends on which rule this is), 20% the
    catalog's flavor text shows. Each active rule this match rolls
    independently.
  - A small, explicit set of `Gameplay`/`Rule-Specific` Toast rows are
    known alternates to specific existing mid-match banners (Combo x4+,
    Fallen Ace, Pollination/Ascension, Smoked Out/Descension, Plus) — see
    GATED_FALLBACKS below.
  - Everything else is NOT gated — it fires 100% of the time whenever its
    trigger condition is true. This includes: Ambiance (idle), Loading
    (always shows one message from whichever rows are currently eligible —
    there's no existing loading-screen system to fall back to), Achievement
    /Milestones (never want to risk missing these — no unlock/page system,
    just a "fire once, at the exact threshold crossing" guard in code), and
    Win Banner/You Lose Banner (flawless outcomes are rare enough already,
    explicitly exempted from the gate).
"""
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
# The repo's own tracked copy is canonical — prefer it. Only fall back to
# Downloads (where earlier edits during this project's original spreadsheet
# session happened to live) if the repo copy is somehow missing.
_REPO_COPY = REPO_ROOT / "tools" / "Honeycomb_Fun_Messages.xlsx"
_DOWNLOADS_COPY = Path.home() / "Downloads" / "Honeycomb_Fun_Messages.xlsx"
DEFAULT_SOURCE = _REPO_COPY if _REPO_COPY.exists() else _DOWNLOADS_COPY

OUT_JSON = REPO_ROOT / "shared/Honeycomb/Resources/HoneycombBannerCatalog.json"
# Windows can't reference the shared/ copy directly (separate build, no shared
# resource-bundling mechanism between the two), so it gets its own copy of the
# same generated content — written here too so one run keeps both in sync,
# rather than relying on a manual copy step after every regeneration.
OUT_JSON_WINDOWS = REPO_ROOT / "windows/src/SoliBee.Desktop/Assets/HoneycombBannerCatalog.json"
OUT_SWIFT = REPO_ROOT / "shared/Honeycomb/Models/BannerID.swift"
OUT_CS = REPO_ROOT / "windows/src/SoliBee.Core/Models/BannerId.cs"

EXPECTED_HEADERS = ["Category", "Trigger", "Message", "Type", "Location", "Spanish"]

TYPE_MAP = {
    "Ambiance": "ambiance",
    "Repeatable Flavor": "repeatableFlavor",
    "Achievement": "achievement",
}
LOCATION_MAP = {
    "Toast": "toast",
    "Loading": "loading",
    "Win Banner": "winBanner",
    "You Lose Banner": "loseBanner",
    "Game intro rules banner": "rulesBanner",
}

GATE_CHANCE = 0.20

# Exact (Category, Trigger) matches that are gated at GATE_CHANCE with a
# fallback to a specific existing production banner. `rulesBanner`-location
# rows are gated automatically elsewhere (dynamic per-rule fallback), so
# they're not listed here.
GATED_FALLBACKS = {
    # Verified against the actual shipped banner text (HoneycombRule.rawValue on mac,
    # HoneycombRule.DisplayName() on Windows — both say the exact same string; see
    # shared/Honeycomb/Models/HoneycombBoard.swift and
    # windows/src/SoliBee.Core/Models/HoneycombTypes.cs). No suit name is appended to
    # the Pollination/Smoked Out banner today — an earlier draft of this dict guessed
    # a suit-annotated fallback ("Ascension: {suit}!") that doesn't match production.
    ("Gameplay", "Combo x4 or higher."): "HIVE MIND x{ComboCount}!",
    ("Rule-Specific", "Fallen Ace triggers (a '1' captures a '10')."): "Queen's Fall!",
    ("Rule-Specific", "Pollination pushes a card's modifier to +3 or higher."): "Pollination!",
    ("Rule-Specific", "Smoked Out drops a card's effective stat to 1."): "Smoked Out!",
    ('Rule-Specific', 'A player triggers a "Plus" combo (the math matches perfectly).'): "Math Bee!",
}

RULE_NAME_FALLBACK_SENTINEL = "$RULE_NAME"


def slugify(text: str, max_len: int = 60) -> str:
    # Strip apostrophes before splitting into words (not just at the edges) so
    # "player's" becomes the one word "players", not two words "player"/"s" —
    # and, critically, never leaves a literal apostrophe embedded in a slug,
    # which would produce an invalid Swift/C# identifier once camel/pascal-cased.
    cleaned = text.replace("'", "").replace("’", "")
    words = re.findall(r"[A-Za-z0-9]+", cleaned.lower())
    slug = ""
    for w in words:
        candidate = f"{slug}_{w}" if slug else w
        if len(candidate) > max_len:
            break
        slug = candidate
    return slug or "unnamed"


def snake_to_lower_camel(snake: str) -> str:
    parts = snake.split("_")
    return parts[0] + "".join(p.capitalize() for p in parts[1:])


def snake_to_pascal(snake: str) -> str:
    return "".join(p.capitalize() for p in snake.split("_"))


def load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected header row {headers!r}; expected {EXPECTED_HEADERS!r}")

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None for v in vals):
            continue
        category, trigger, message, type_, location, spanish = vals
        # Spanish is the one column allowed to be blank — translation lands
        # incrementally, so an untranslated row must not block regeneration.
        # Every other column is still required.
        missing = [
            name for name, v in zip(EXPECTED_HEADERS, vals)
            if v is None and name != "Spanish"
        ]
        if missing:
            raise ValueError(f"Row {r} is missing {missing}: {vals!r}")
        if type_.strip() not in TYPE_MAP:
            raise ValueError(f"Row {r}: unknown Type {type_!r}")
        if location.strip() not in LOCATION_MAP:
            raise ValueError(f"Row {r}: unknown Location {location!r}")
        # Both a blank cell and the literal "No Translation" (the translator's
        # explicit "deliberately skipped this one" marker) normalize to "" —
        # the runtime treats "" as "exclude this message from the Spanish
        # pool entirely" (see BannerCatalog.pickMessage), not as an English
        # fallback for that one message.
        spanish_text = spanish.strip() if spanish else ""
        if spanish_text.lower() == "no translation":
            spanish_text = ""
        rows.append({
            "category": category.strip(),
            "trigger": trigger.strip(),
            "message": message.strip(),
            "type": type_.strip(),
            "location": location.strip(),
            "spanish": spanish_text,
        })
    return rows


def group_rows(rows: list[dict]) -> list[dict]:
    groups: "OrderedDict[tuple[str, str], dict]" = OrderedDict()
    for row in rows:
        key = (row["category"], row["trigger"])
        if key not in groups:
            groups[key] = {
                "category": row["category"],
                "trigger": row["trigger"],
                "type": row["type"],
                "location": row["location"],
                "messages": [],
                "messages_es": [],
            }
        g = groups[key]
        if g["type"] != row["type"] or g["location"] != row["location"]:
            raise ValueError(
                f"Inconsistent Type/Location within group {key}: "
                f"{g['type']}/{g['location']} vs {row['type']}/{row['location']}"
            )
        if row["message"] in g["messages"]:
            raise ValueError(f"Duplicate message within group {key}: {row['message']!r}")
        g["messages"].append(row["message"])
        g["messages_es"].append(row["spanish"])
    return list(groups.values())


def build_catalog(groups: list[dict]) -> list[dict]:
    catalog = []
    used_ids: set[str] = set()
    for g in groups:
        base_id = f"{slugify(g['category'])}_{slugify(g['trigger'])}"
        entry_id = base_id
        n = 2
        while entry_id in used_ids:
            entry_id = f"{base_id}_{n}"
            n += 1
        used_ids.add(entry_id)

        type_key = TYPE_MAP[g["type"]]
        location_key = LOCATION_MAP[g["location"]]
        key = (g["category"], g["trigger"])

        gated = False
        fallback = None
        if location_key == "rulesBanner":
            gated = True
            fallback = RULE_NAME_FALLBACK_SENTINEL
        elif key in GATED_FALLBACKS:
            gated = True
            fallback = GATED_FALLBACKS[key]

        catalog.append({
            "id": entry_id,
            "category": g["category"],
            "trigger": g["trigger"],
            "type": type_key,
            "location": location_key,
            "gated": gated,
            "gateChance": GATE_CHANCE if gated else None,
            "fallback": fallback,
            # `fallback` text (existing production banner strings like "Queen's
            # Fall!", or the rule-name sentinel) is NOT translated here — those
            # strings mirror other already-shipped English text (e.g.
            # HoneycombRule display names) that isn't itself part of either
            # localization pipeline yet. Translating just the fallback would
            # create a second, inconsistent source of truth for that text.
            # Tracked as a known follow-up, not silently done here.
            "messages": g["messages"],
            "messagesEs": g["messages_es"],
        })
    return catalog


def write_json(catalog: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = {
        "version": 1,
        "generatedBy": "tools/generate_banner_catalog.py — do not hand-edit",
        "banners": catalog,
    }
    path.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_swift(catalog: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_banner_catalog.py` from",
        "// Honeycomb_Fun_Messages.xlsx. See that script for the id/gating rules.",
        "",
        "// Stable identifier for every banner/toast catalog entry (mac + iOS).",
        "// The raw value is the JSON catalog's `id` field — HoneycombBannerCatalog.json",
        "// is keyed by this exact string, so it must never be renamed or reused once",
        "// shipped (an achievement/milestone's \"already fired\" state, if ever",
        "// persisted, would key off this too).",
        "public enum BannerID: String, CaseIterable, Codable {",
    ]
    for entry in catalog:
        case_name = snake_to_lower_camel(entry["id"])
        lines.append(f'    case {case_name} = "{entry["id"]}"')
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cs(catalog: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// GENERATED FILE — do not hand-edit.",
        "// Regenerate via `python3 tools/generate_banner_catalog.py` from",
        "// Honeycomb_Fun_Messages.xlsx. See that script for the id/gating rules.",
        "",
        "using System.Collections.Generic;",
        "",
        "namespace SoliBee.Core.Models;",
        "",
        "// Stable identifier for every banner/toast catalog entry (Windows). Mirrors",
        "// the Swift port's BannerID (shared/Honeycomb/Models/BannerID.swift) —",
        "// same catalog, same ids, generated from the same spreadsheet in one pass.",
        "public enum BannerId",
        "{",
    ]
    pascal_names = [snake_to_pascal(entry["id"]) for entry in catalog]
    for name in pascal_names:
        lines.append(f"    {name},")
    lines.append("}")
    lines.append("")
    lines.append("public static class BannerIdExtensions")
    lines.append("{")
    lines.append("    // Maps the JSON catalog's snake_case `id` string to its BannerId —")
    lines.append("    // needed because System.Text.Json won't auto-match snake_case ids to")
    lines.append("    // PascalCase enum members.")
    lines.append("    private static readonly Dictionary<string, BannerId> ById = new()")
    lines.append("    {")
    for entry, name in zip(catalog, pascal_names):
        lines.append(f'        ["{entry["id"]}"] = BannerId.{name},')
    lines.append("    };")
    lines.append("")
    lines.append("    public static BannerId Parse(string id) => ById[id];")
    lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        print(f"Source spreadsheet not found: {source}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(source)
    groups = group_rows(rows)
    catalog = build_catalog(groups)

    write_json(catalog, OUT_JSON)
    write_json(catalog, OUT_JSON_WINDOWS)
    write_swift(catalog, OUT_SWIFT)
    write_cs(catalog, OUT_CS)

    gated_count = sum(1 for e in catalog if e["gated"])
    print(f"Read {len(rows)} spreadsheet rows -> {len(catalog)} catalog entries "
          f"({gated_count} gated, {len(catalog) - gated_count} always-fire).")
    for out in (OUT_JSON, OUT_JSON_WINDOWS, OUT_SWIFT, OUT_CS):
        print(f"Wrote {out.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
